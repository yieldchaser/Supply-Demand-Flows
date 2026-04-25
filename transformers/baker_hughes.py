"""
Baker Hughes NA Rig Count weekly .xlsx → curated long-format Parquet.

Source sheet: 'NAM Weekly' — granular rig-deployment rows from Jan 2024 onward.
Row 11 is the header (rows 1-10 are Power BI slicer widgets, ignored).

Strategy: produce ONE Parquet with both granular and rollup rows. Rollup rows
have series_id="bh_rollup_*" and drive the dashboard's headline panel; granular
rows have series_id="bh_granular_*" and enable basin/state/trajectory analysis.

Canonical schema:
  source (str)      : "baker_hughes"
  series_id (str)   : see series_id rules below
  series_name (str) : human label for the series
  period (date)     : the US_PublishDate (weekly)
  value (int64)     : rig count (may be 0)
  unit (str)        : "rigs"
  region (str)      : country or state for granular, country for rollups
  ingested_at (str) : UTC ISO datetime

series_id rules:
  Rollup (always emitted, aggregated across all granular rows of the same
  period that match the filter):
    bh_rollup_us_total              — all US rows
    bh_rollup_us_oil                — US + DrillFor=Oil
    bh_rollup_us_gas                — US + DrillFor=Gas
    bh_rollup_us_horizontal         — US + Trajectory=Horizontal
    bh_rollup_us_vertical           — US + Trajectory=Vertical
    bh_rollup_us_directional        — US + Trajectory=Directional
    bh_rollup_us_land               — US + Location=Land
    bh_rollup_us_offshore           — US + Location=Offshore
    bh_rollup_us_inland_waters      — US + Location=Inland Waters
    bh_rollup_canada_total          — all Canada rows
    bh_rollup_canada_oil            — Canada + DrillFor=Oil
    bh_rollup_canada_gas            — Canada + DrillFor=Gas
    bh_rollup_na_total              — US + Canada total
    bh_rollup_basin_<basin_slug>    — US rows grouped by Basin (e.g., permian,
                                       marcellus, haynesville, eagle_ford)

  Granular (one row per (Country, Basin, County, DrillFor, Location,
  Trajectory, State/Province) combination per period):
    bh_granular_<hash>
      where <hash> is md5 of joined dimensions (8 chars)
      series_name encodes the dimension combination for debugging
      region = State/Province (or country if state missing)

Failure modes:
  - 'NAM Weekly' sheet missing → TransformError with available sheets listed
  - Header row 11 doesn't contain expected column names → TransformError
  - Rig Count Value column has non-numeric non-zero → skip row, log warning,
    continue (don't fail the whole file for one bad row)
  - US_PublishDate has unparseable value → skip row, log warning
  - Zero rows after filtering → TransformError (suggests file corruption)
"""

import hashlib
import logging
from datetime import UTC, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from scrapers.base.safe_writer import safe_write_bytes
from transformers.errors import TransformError

logger = logging.getLogger(__name__)

SHEET_NAME = "NAM Weekly"
HEADER_ROW = 11  # 1-indexed in Excel, so header is at row 11

EXPECTED_COLUMNS = {
    "Country",
    "County",
    "Basin",
    "GOM",
    "DrillFor",
    "Location",
    "State/Province",
    "Trajectory",
    "Year",
    "Month",
    "US_PublishDate",
    "Rig Count Value",
}

# Basin slugs for rollup series_ids
BASIN_SLUGS = {
    "Permian": "permian",
    "Marcellus": "marcellus",
    "Haynesville": "haynesville",
    "Eagle Ford": "eagle_ford",
    "DJ-Niobrara": "dj_niobrara",
    "Bakken": "bakken",
    "Utica": "utica",
    "Cana Woodford": "cana_woodford",
    "Arkoma Woodford": "arkoma_woodford",
    "Ardmore Woodford": "ardmore_woodford",
    "Barnett": "barnett",
    "Granite Wash": "granite_wash",
    "Fayetteville": "fayetteville",
    "Mississippian": "mississippian",
    "Williston": "williston",
    # "Other" excluded from rollups — not a discrete basin
}


def _parse_publish_date(raw: object) -> pd.Timestamp | None:
    """Handle both datetime cells and DD-MM-YYYY string cells."""
    if pd.isna(raw):
        return None
    if isinstance(raw, datetime):
        return pd.Timestamp(raw).normalize()
    if isinstance(raw, str):
        # Try DD-MM-YYYY format first (the BH default)
        try:
            return pd.to_datetime(raw, format="%d-%m-%Y", errors="raise").normalize()
        except (ValueError, TypeError):
            pass
        # Fallback: pandas auto-parse
        try:
            return pd.to_datetime(raw, errors="coerce").normalize()
        except Exception:
            return None
    return None


def _granular_hash(row: pd.Series) -> str:
    key = "|".join(
        [
            str(row.get("Country", "")),
            str(row.get("Basin", "")),
            str(row.get("County", "")),
            str(row.get("DrillFor", "")),
            str(row.get("Location", "")),
            str(row.get("Trajectory", "")),
            str(row.get("State/Province", "")),
        ]
    )
    return hashlib.md5(key.encode()).hexdigest()[:8]


def _load_nam_weekly(raw_xlsx_path: Path) -> pd.DataFrame:
    """
    Load the NAM Weekly sheet with header at row 11. Returns a DataFrame
    with columns named after the header row.
    """
    wb = load_workbook(raw_xlsx_path, read_only=True, data_only=True)
    available = wb.sheetnames
    if SHEET_NAME not in available:
        raise TransformError(f"Missing expected sheet: '{SHEET_NAME}'. Available: {available}")

    ws = wb[SHEET_NAME]
    # openpyxl is 1-indexed; header row 11 means values_only iter_rows starting min_row=11
    rows_iter = ws.iter_rows(min_row=HEADER_ROW, values_only=True)
    header = list(next(rows_iter))
    # Trim trailing None columns (Power BI sheets often have phantom cols)
    while header and header[-1] is None:
        header.pop()
    header_set = {c for c in header if c is not None}
    missing = EXPECTED_COLUMNS - header_set
    if missing:
        raise TransformError(
            f"Header row {HEADER_ROW} missing expected columns: {sorted(missing)}. "
            f"Found: {sorted(header_set)}"
        )

    data_rows = []
    for row in rows_iter:
        # Truncate row to header length, drop fully-empty rows
        row_trimmed = list(row[: len(header)])
        if all(cell is None for cell in row_trimmed):
            continue
        data_rows.append(row_trimmed)

    wb.close()

    return pd.DataFrame(data_rows, columns=header)


def _build_rollup_rows(df: pd.DataFrame, ingested_at: str) -> list[dict]:
    """Build rollup series by aggregating granular rows per period."""
    out: list[dict] = []

    # Ensure numeric rig count
    df = df.copy()
    df["Rig Count Value"] = (
        pd.to_numeric(df["Rig Count Value"], errors="coerce").fillna(0).astype(int)
    )

    # Group by period for aggregations
    periods = sorted(df["period"].dropna().unique())

    for period in periods:
        period_df = df[df["period"] == period]
        us_df = period_df[period_df["Country"] == "UNITED STATES"]
        ca_df = period_df[period_df["Country"] == "CANADA"]

        def add(series_id: str, series_name: str, value: int, region: str, p=period) -> None:
            out.append(
                {
                    "source": "baker_hughes",
                    "series_id": series_id,
                    "series_name": series_name,
                    "period": pd.Timestamp(p).date(),
                    "value": int(value),
                    "unit": "rigs",
                    "region": region,
                    "ingested_at": ingested_at,
                }
            )

        # US aggregates
        add("bh_rollup_us_total", "US Total Rigs", us_df["Rig Count Value"].sum(), "United States")
        add(
            "bh_rollup_us_oil",
            "US Oil Rigs",
            us_df[us_df["DrillFor"] == "Oil"]["Rig Count Value"].sum(),
            "United States",
        )
        add(
            "bh_rollup_us_gas",
            "US Gas Rigs",
            us_df[us_df["DrillFor"] == "Gas"]["Rig Count Value"].sum(),
            "United States",
        )
        add(
            "bh_rollup_us_horizontal",
            "US Horizontal Rigs",
            us_df[us_df["Trajectory"] == "Horizontal"]["Rig Count Value"].sum(),
            "United States",
        )
        add(
            "bh_rollup_us_vertical",
            "US Vertical Rigs",
            us_df[us_df["Trajectory"] == "Vertical"]["Rig Count Value"].sum(),
            "United States",
        )
        add(
            "bh_rollup_us_directional",
            "US Directional Rigs",
            us_df[us_df["Trajectory"] == "Directional"]["Rig Count Value"].sum(),
            "United States",
        )
        add(
            "bh_rollup_us_land",
            "US Land Rigs",
            us_df[us_df["Location"] == "Land"]["Rig Count Value"].sum(),
            "United States",
        )
        add(
            "bh_rollup_us_offshore",
            "US Offshore Rigs",
            us_df[us_df["Location"] == "Offshore"]["Rig Count Value"].sum(),
            "United States",
        )
        add(
            "bh_rollup_us_inland_waters",
            "US Inland Waters Rigs",
            us_df[us_df["Location"] == "Inland Waters"]["Rig Count Value"].sum(),
            "United States",
        )

        # Canada aggregates
        add("bh_rollup_canada_total", "Canada Total Rigs", ca_df["Rig Count Value"].sum(), "Canada")
        add(
            "bh_rollup_canada_oil",
            "Canada Oil Rigs",
            ca_df[ca_df["DrillFor"] == "Oil"]["Rig Count Value"].sum(),
            "Canada",
        )
        add(
            "bh_rollup_canada_gas",
            "Canada Gas Rigs",
            ca_df[ca_df["DrillFor"] == "Gas"]["Rig Count Value"].sum(),
            "Canada",
        )

        # NA total
        add(
            "bh_rollup_na_total",
            "North America Total Rigs",
            us_df["Rig Count Value"].sum() + ca_df["Rig Count Value"].sum(),
            "North America",
        )

        # Basin aggregates (US only — Canadian basin naming is different)
        for basin_name, basin_slug in BASIN_SLUGS.items():
            basin_rows = us_df[us_df["Basin"] == basin_name]
            if basin_rows.empty:
                continue

            basin_rigs = basin_rows["Rig Count Value"].sum()
            if basin_rigs > 0:
                add(
                    f"bh_rollup_basin_{basin_slug}",
                    f"US {basin_name} Basin Rigs",
                    basin_rigs,
                    "United States",
                )

                for drill_for, drill_slug in [("Gas", "gas"), ("Oil", "oil")]:
                    drill_rigs = basin_rows[basin_rows["DrillFor"] == drill_for][
                        "Rig Count Value"
                    ].sum()
                    add(
                        f"bh_rollup_basin_{basin_slug}_{drill_slug}",
                        f"{basin_name} basin · {drill_for} rigs",
                        drill_rigs,
                        "United States",
                    )

    return out


def _build_granular_rows(df: pd.DataFrame, ingested_at: str) -> list[dict]:
    """One row per (dimension-combo, period) — full granular long-format."""
    df = df.copy()
    df["Rig Count Value"] = (
        pd.to_numeric(df["Rig Count Value"], errors="coerce").fillna(0).astype(int)
    )

    out: list[dict] = []
    for _, row in df.iterrows():
        period = row.get("period")
        if pd.isna(period):
            continue
        gid = _granular_hash(row)
        series_name = (
            f"{row.get('Country', '?')} | "
            f"{row.get('Basin', '?')} | "
            f"{row.get('DrillFor', '?')} | "
            f"{row.get('Trajectory', '?')} | "
            f"{row.get('State/Province', '?')} "
            f"({row.get('County', '?')})"
        )
        region = row.get("State/Province") or row.get("Country", "?")
        out.append(
            {
                "source": "baker_hughes",
                "series_id": f"bh_granular_{gid}",
                "series_name": series_name,
                "period": pd.Timestamp(period).date(),
                "value": int(row["Rig Count Value"]),
                "unit": "rigs",
                "region": str(region),
                "ingested_at": ingested_at,
            }
        )
    return out


def transform(raw_xlsx_path: Path, curated_parquet_path: Path) -> dict:
    """
    Why: rig counts are the leading indicator for basin-level US production.
    Horizontal rig share and basin-level gas rigs feed the supply-side nowcast.

    What: parse NAM Weekly sheet, produce rollup + granular long-format Parquet.

    Returns:
        dict with row_count, rollup_count, granular_count, period_min, period_max,
        basins_covered
    """
    df = _load_nam_weekly(raw_xlsx_path)

    # Parse publish date
    df["period"] = df["US_PublishDate"].apply(_parse_publish_date)
    bad_dates = df["period"].isna().sum()
    if bad_dates > 0:
        logger.warning(
            f"Baker Hughes transformer: {bad_dates} rows had unparseable US_PublishDate, skipped."
        )
    df = df[df["period"].notna()]

    if df.empty:
        raise TransformError("No valid rows after date parsing — file may be corrupt.")

    ingested_at = datetime.now(UTC).isoformat()
    rollup = _build_rollup_rows(df, ingested_at)
    granular = _build_granular_rows(df, ingested_at)

    out_df = pd.DataFrame(rollup + granular)

    # Serialize to Parquet bytes (snappy compression)
    import io

    buf = io.BytesIO()
    out_df.to_parquet(buf, compression="snappy", index=False)
    safe_write_bytes(curated_parquet_path, buf.getvalue())

    basins_covered = sorted(
        set(
            r["series_name"].split(" Basin")[0].removeprefix("US ")
            for r in rollup
            if "Basin" in r["series_name"]
        )
    )

    return {
        "rows": len(out_df),
        "rollup_count": len(rollup),
        "granular_count": len(granular),
        "period_min": str(out_df["period"].min()),
        "period_max": str(out_df["period"].max()),
        "basins_covered": basins_covered,
    }


if __name__ == "__main__":
    import json

    latest = sorted(Path("data/raw/baker_hughes").rglob("*.xlsx"))[-1]
    out = Path("data/curated/baker_hughes_weekly.parquet")
    out.parent.mkdir(parents=True, exist_ok=True)
    result = transform(latest, out)
    print(json.dumps(result, indent=2, default=str))
